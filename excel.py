from openpyxl import workbook, load_workbook
import random
wb = load_workbook("hello.xlsx")
ws = wb.active
rangeline = ws["A2": "A19"]
name = []
for items in rangeline:
    for subitems in items:
        name.append(subitems.value)

computer_action = random.choice(name)
print("The computer randomly chose: " + computer_action)

